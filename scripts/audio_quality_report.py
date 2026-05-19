"""
Simple audio quality assessment — point at a folder, get an Excel report.

Walks the folder recursively for audio files, computes acoustic metrics per
file, classifies quality, writes a single Excel workbook with color-coded
results.

No setup, no JSON intermediate, no language flag, no mic-name mapping required.

Usage:
    # Basic — per-file quality only (fast, no diarization)
    uv run python scripts/audio_quality_report.py /path/to/audio/folder

    # With per-speaker breakdown (adds a second sheet, slower, needs HF token)
    uv run python scripts/audio_quality_report.py /path/to/audio --with-speakers

    # Custom output location
    uv run python scripts/audio_quality_report.py /path/to/audio \
        -o ./my_quality_report.xlsx

    # Restrict to certain extensions
    uv run python scripts/audio_quality_report.py /path/to/audio \
        --extensions .m4a .wav

Output:
    quality_report.xlsx in the input folder (or wherever -o points)
    Sheet 1 "Per file":
        filename, duration, SNR, RMS, clipping %, speech %, quality, score, comment
    Sheet 2 "Per speaker" (when --with-speakers):
        filename, speaker_id, role (ENUMERATOR / RESPONDENT / OTHER),
        talk_time, n_turns, avg_turn_s, SNR, RMS, quality, score, comment
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

import numpy as np


# ─── Audio loading ───────────────────────────────────────────────────────

def load_audio_mono16k(path: Path) -> tuple[np.ndarray, int]:
    """Load any audio file as mono float32 at 16 kHz. Uses librosa."""
    import librosa
    wav, sr = librosa.load(str(path), sr=16000, mono=True)
    return wav.astype(np.float32), sr


# ─── Quality metrics ─────────────────────────────────────────────────────

def compute_metrics(wav: np.ndarray, sr: int = 16000) -> dict:
    """Compute per-file acoustic quality metrics.
    Returns dict with keys: duration_s, snr_db, rms_dbfs, clipping_pct,
                            speech_pct, peak_dbfs.
    """
    if len(wav) == 0:
        return {
            "duration_s": 0.0, "snr_db": 0.0, "rms_dbfs": -120.0,
            "clipping_pct": 0.0, "speech_pct": 0.0, "peak_dbfs": -120.0,
        }

    duration_s = len(wav) / sr

    # RMS (loudness)
    rms = np.sqrt(np.mean(wav.astype(np.float64) ** 2))
    rms_dbfs = 20 * np.log10(max(rms, 1e-12))

    # Peak level
    peak = np.max(np.abs(wav))
    peak_dbfs = 20 * np.log10(max(peak, 1e-12))

    # Clipping detection (samples within 0.5dB of full scale)
    clipping_threshold = 0.997
    clipping_pct = float(np.mean(np.abs(wav) >= clipping_threshold) * 100)

    # SNR via frame-energy percentile method
    # Frame the signal at 20ms hops
    frame_samples = int(0.02 * sr)
    n_frames = len(wav) // frame_samples
    if n_frames < 5:
        snr_db = 0.0
        speech_pct = 0.0
    else:
        frames = wav[: n_frames * frame_samples].reshape(n_frames, frame_samples)
        frame_energy = np.sqrt(np.mean(frames.astype(np.float64) ** 2, axis=1))
        # Avoid log(0)
        frame_energy = np.maximum(frame_energy, 1e-9)
        frame_db = 20 * np.log10(frame_energy)

        # Signal = upper 25th percentile, noise = lower 25th percentile
        signal_db = np.percentile(frame_db, 75)
        noise_db = np.percentile(frame_db, 25)
        snr_db = float(signal_db - noise_db)

        # Speech % = frames significantly above noise floor (15 dB margin)
        speech_threshold_db = noise_db + 15
        speech_pct = float(np.mean(frame_db > speech_threshold_db) * 100)

    return {
        "duration_s": float(duration_s),
        "snr_db": float(snr_db),
        "rms_dbfs": float(rms_dbfs),
        "clipping_pct": float(clipping_pct),
        "speech_pct": float(speech_pct),
        "peak_dbfs": float(peak_dbfs),
    }


# ─── Quality classification ──────────────────────────────────────────────

def classify_quality(snr: float, speech_pct: float,
                     clipping_pct: float, rms_dbfs: float,
                     duration_s: float = 0.0) -> tuple[str, str, int]:
    """Return (flag, comment, score 0-100).

    Thresholds tuned for phone-mic field interview audio. Hierarchical:
    technical issues (clipping, muted, too short) take precedence over
    content quality (SNR, speech %).
    """
    # ─── Technical issues (override everything else) ────────────────
    if duration_s > 0 and duration_s < 1.0:
        return "BROKEN", f"too short ({duration_s:.1f}s) — likely incomplete recording", 0

    if clipping_pct > 1.0:
        sev = min(100, int(clipping_pct * 30))
        return "CLIPPED", \
               f"heavy distortion: {clipping_pct:.2f}% samples clipped — reduce mic gain", \
               max(10, 40 - sev)

    if clipping_pct > 0.1:
        return "CLIPPED", \
               f"some distortion: {clipping_pct:.2f}% clipped — slightly hot recording", \
               60

    if rms_dbfs < -55:
        return "MUTED", "no signal detected — mic likely off or unplugged", 0

    if rms_dbfs < -45:
        return "BAD", f"very quiet (RMS {rms_dbfs:.0f} dBFS) — mic likely muted or far away", 15

    # ─── Content quality based on SNR + speech % ───────────────────
    if snr < 6:
        return "BAD", f"barely audible (SNR {snr:.0f} dB) — no usable speech", 10

    if snr < 10:
        return "POOR", f"very noisy (SNR {snr:.0f} dB) — ASR will struggle, high WER expected", 25

    if snr < 13:
        return "POOR", f"noisy (SNR {snr:.0f} dB) — usable but expect errors", 40

    # SNR is decent — refine by speech %
    if speech_pct < 10:
        return "EMPTY", \
               f"almost no speech ({speech_pct:.0f}%) — possibly background recording", 30

    if speech_pct < 25:
        return "LOW SPEECH", \
               f"clean audio (SNR {snr:.0f} dB) but mostly silent ({speech_pct:.0f}% speech) — mic far/muted speaker", 55

    if snr >= 25 and speech_pct >= 50 and clipping_pct < 0.05:
        return "EXCELLENT", \
               f"studio-quality (SNR {snr:.0f} dB, {speech_pct:.0f}% speech) — ideal for ASR", 95

    if snr >= 20 and speech_pct >= 35:
        return "GOOD", \
               f"clean, speech-rich (SNR {snr:.0f} dB, {speech_pct:.0f}% speech)", 85

    if snr >= 15 and speech_pct >= 25:
        return "FAIR", \
               f"acceptable (SNR {snr:.0f} dB, {speech_pct:.0f}% speech) — expect minor errors", 65

    return "FAIR", \
           f"marginal (SNR {snr:.0f} dB, {speech_pct:.0f}% speech) — review recommended", 50


# ─── Excel writer ────────────────────────────────────────────────────────

# ─── Diarization (optional) ──────────────────────────────────────────────

def load_hf_token():
    """Try to find HF_TOKEN: environment, then .env files."""
    import os
    token = os.environ.get("HF_TOKEN") or os.environ.get("HUGGING_FACE_HUB_TOKEN")
    if token:
        return token
    TOKEN_KEYS = {
        "HF_TOKEN", "HUGGING_FACE_HUB_TOKEN", "HUGGINGFACE_TOKEN",
        "HF_HUB_TOKEN", "HUGGINGFACE_HUB_TOKEN",
    }
    for env_path in [Path(".env"), Path(__file__).parent.parent / ".env",
                     Path.home() / ".env"]:
        if not env_path.exists():
            continue
        with open(env_path, encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#") or "=" not in line:
                    continue
                key, _, val = line.partition("=")
                if key.strip().upper() in TOKEN_KEYS:
                    return val.strip().strip('"').strip("'")
    return None


def load_diarizer():
    """Load pyannote 3.1 speaker diarization pipeline. Returns None on failure."""
    import torch
    try:
        from pyannote.audio import Pipeline
    except ImportError:
        print("ERROR: pyannote.audio not installed.")
        print("Run: uv pip install pyannote.audio")
        return None

    token = load_hf_token()
    if not token:
        print("ERROR: No HF_TOKEN found.")
        print("Set HF_TOKEN in .env or env, and accept terms at:")
        print("  https://huggingface.co/pyannote/speaker-diarization-3.1")
        return None

    # PyTorch 2.6+ defaults torch.load to weights_only=True, which rejects
    # pyannote's checkpoint. Force weights_only=False for trusted HF repos.
    _orig_load = torch.load
    def _patched_load(*a, **kw):
        kw["weights_only"] = False  # Force override, even if kw has weights_only=None
        return _orig_load(*a, **kw)
    torch.load = _patched_load

    try:
        pipeline = Pipeline.from_pretrained(
            "pyannote/speaker-diarization-3.1",
            use_auth_token=token,
        )
    except Exception as e:
        print(f"ERROR loading diarization pipeline: {e}")
        print("Make sure you accepted terms at:")
        print("  https://huggingface.co/pyannote/speaker-diarization-3.1")
        torch.load = _orig_load
        return None
    finally:
        torch.load = _orig_load

    if torch.cuda.is_available():
        pipeline = pipeline.to(torch.device("cuda"))
        print("  Diarizer on GPU")
    else:
        print("  Diarizer on CPU (slower)")
    return pipeline


def diarize_file(diarizer, wav: np.ndarray, sr: int = 16000):
    """Run diarization on in-memory audio. Returns dict of
       speaker_id -> list of (start_s, end_s) tuples."""
    import torch
    from collections import defaultdict
    audio_tensor = torch.from_numpy(wav).float().unsqueeze(0)
    diarization = diarizer({"waveform": audio_tensor, "sample_rate": sr})

    speaker_segments = defaultdict(list)
    for segment, _, speaker_id in diarization.itertracks(yield_label=True):
        speaker_segments[speaker_id].append((segment.start, segment.end))
    return dict(speaker_segments)


def extract_speaker_audio(wav: np.ndarray, segments: list, sr: int = 16000) -> np.ndarray:
    """Concatenate the speaker's segments into one continuous waveform."""
    if not segments:
        return np.zeros(0, dtype=np.float32)
    parts = []
    for start_s, end_s in segments:
        a = int(start_s * sr)
        b = int(end_s * sr)
        a = max(0, min(a, len(wav)))
        b = max(0, min(b, len(wav)))
        if b > a:
            parts.append(wav[a:b])
    return np.concatenate(parts) if parts else np.zeros(0, dtype=np.float32)


def label_speaker_roles(per_speaker_stats: dict) -> dict[str, str]:
    """Heuristic: ENUMERATOR speaks first AND has more / shorter turns."""
    if not per_speaker_stats:
        return {}
    ranked = sorted(per_speaker_stats.values(),
                    key=lambda s: s["talk_time_s"], reverse=True)
    if len(ranked) == 1:
        return {ranked[0]["speaker_id"]: "UNKNOWN"}

    a, b = ranked[0], ranked[1]
    extras = ranked[2:]

    def score(cand, other):
        s = 0.0
        if cand["first_speaks_at_s"] < other["first_speaks_at_s"]:
            s += 2.0
        if cand["n_turns"] > other["n_turns"]:
            s += 1.0
        if cand["avg_turn_s"] < other["avg_turn_s"]:
            s += 1.0
        return s

    a_score, b_score = score(a, b), score(b, a)
    if a_score >= b_score:
        roles = {a["speaker_id"]: "ENUMERATOR", b["speaker_id"]: "RESPONDENT"}
    else:
        roles = {a["speaker_id"]: "RESPONDENT", b["speaker_id"]: "ENUMERATOR"}
    for x in extras:
        roles[x["speaker_id"]] = "OTHER"
    return roles


def per_speaker_analysis(wav: np.ndarray, diarizer, sr: int = 16000) -> list[dict]:
    """Return list of per-speaker rows for one audio file (or [] if diar fails)."""
    if diarizer is None or len(wav) == 0:
        return []
    try:
        speaker_segments = diarize_file(diarizer, wav, sr)
    except Exception as e:
        print(f"      diarization error: {e}")
        return []
    if not speaker_segments:
        return []

    # Compute per-speaker stats
    stats = {}
    for spk_id, segs in speaker_segments.items():
        if not segs:
            continue
        durations = [end - start for start, end in segs]
        stats[spk_id] = {
            "speaker_id": spk_id,
            "first_speaks_at_s": min(start for start, _ in segs),
            "talk_time_s": sum(durations),
            "n_turns": len(segs),
            "avg_turn_s": sum(durations) / len(durations),
            "segments": segs,
        }

    roles = label_speaker_roles(stats)

    rows = []
    for spk_id, s in stats.items():
        spk_audio = extract_speaker_audio(wav, s["segments"], sr)
        if len(spk_audio) < sr * 0.5:    # < 0.5s of speech, skip
            continue
        metrics = compute_metrics(spk_audio, sr)
        flag, comment, score = classify_quality(
            metrics["snr_db"], metrics["speech_pct"],
            metrics["clipping_pct"], metrics["rms_dbfs"],
            metrics["duration_s"],
        )
        rows.append({
            "speaker_id": spk_id,
            "role": roles.get(spk_id, "UNKNOWN"),
            "talk_time_s": round(s["talk_time_s"], 1),
            "n_turns": s["n_turns"],
            "avg_turn_s": round(s["avg_turn_s"], 1),
            "first_speaks_at_s": round(s["first_speaks_at_s"], 1),
            **{f"speaker_{k}": v for k, v in metrics.items()},
            "quality": flag,
            "score": score,
            "comment": comment,
        })

    # Sort by role priority then talk time
    role_order = {"ENUMERATOR": 0, "RESPONDENT": 1, "OTHER": 2, "UNKNOWN": 3}
    rows.sort(key=lambda r: (role_order.get(r["role"], 9), -r["talk_time_s"]))
    return rows


# ─── Excel writer ────────────────────────────────────────────────────────

def write_excel(rows: list[dict], out_path: Path,
                per_speaker_rows: list[dict] | None = None):
    """Write quality report to Excel with color-coded quality column."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("ERROR: openpyxl not installed. Run: uv pip install openpyxl")
        sys.exit(1)

    wb = Workbook()
    ws = wb.active
    ws.title = "Audio Quality"

    columns = [
        ("file", 50),
        ("folder", 30),
        ("duration_s", 12),
        ("snr_db", 10),
        ("rms_dbfs", 10),
        ("peak_dbfs", 11),
        ("clipping_pct", 13),
        ("speech_pct", 12),
        ("quality", 13),
        ("score", 8),
        ("comment", 60),
    ]
    headers = [c[0] for c in columns]
    ws.append(headers)
    for i, h in enumerate(headers, 1):
        ws.cell(1, i).font = Font(bold=True)
        ws.cell(1, i).alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(i)].width = columns[i-1][1]
    ws.freeze_panes = "A2"

    # Color map
    quality_colors = {
        "EXCELLENT":  "63BE7B",  # bright green
        "GOOD":       "C6EFCE",  # light green
        "FAIR":       "FFEB9C",  # light yellow
        "LOW SPEECH": "FFD18C",  # orange
        "EMPTY":      "F4B084",  # darker orange
        "POOR":       "FFC7CE",  # light red
        "BAD":        "FF6B6B",  # red
        "CLIPPED":    "B19CD9",  # purple
        "MUTED":      "808080",  # gray
        "BROKEN":     "404040",  # dark gray
    }

    for r in rows:
        ws.append([
            r["file"],
            r["folder"],
            round(r["duration_s"], 2),
            round(r["snr_db"], 1),
            round(r["rms_dbfs"], 1),
            round(r["peak_dbfs"], 1),
            round(r["clipping_pct"], 3),
            round(r["speech_pct"], 1),
            r["quality"],
            r.get("score", 0),
            r["comment"],
        ])
        # Color the quality cell
        quality_col_idx = headers.index("quality") + 1
        cell = ws.cell(row=ws.max_row, column=quality_col_idx)
        color = quality_colors.get(r["quality"])
        if color:
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

    # Summary footer
    n = len(rows)
    summary_row = ws.max_row + 2
    ws.cell(summary_row, 1).value = "SUMMARY"
    ws.cell(summary_row, 1).font = Font(bold=True, size=12)

    summary_row += 1
    ws.cell(summary_row, 1).value = "Total files"
    ws.cell(summary_row, 2).value = n
    ws.cell(summary_row, 1).font = Font(bold=True)

    # Quality distribution
    from collections import Counter
    quality_counts = Counter(r["quality"] for r in rows)
    summary_row += 1
    ws.cell(summary_row, 1).value = "Quality distribution:"
    ws.cell(summary_row, 1).font = Font(bold=True)
    for flag in ["EXCELLENT", "GOOD", "FAIR", "LOW SPEECH", "EMPTY",
                 "POOR", "BAD", "CLIPPED", "MUTED", "BROKEN"]:
        count = quality_counts.get(flag, 0)
        if count == 0:
            continue
        summary_row += 1
        ws.cell(summary_row, 1).value = f"  {flag}"
        ws.cell(summary_row, 2).value = count
        ws.cell(summary_row, 3).value = f"{100*count/n:.0f}%"
        color = quality_colors.get(flag)
        if color:
            ws.cell(summary_row, 1).fill = PatternFill(
                start_color=color, end_color=color, fill_type="solid")

    # Mean metrics
    if n > 0:
        summary_row += 2
        ws.cell(summary_row, 1).value = "Average metrics:"
        ws.cell(summary_row, 1).font = Font(bold=True)
        for label, key in [("Mean SNR (dB)", "snr_db"),
                           ("Mean RMS (dBFS)", "rms_dbfs"),
                           ("Mean speech %", "speech_pct"),
                           ("Mean clipping %", "clipping_pct"),
                           ("Total duration (min)", None)]:
            summary_row += 1
            ws.cell(summary_row, 1).value = f"  {label}"
            if key:
                vals = [r[key] for r in rows]
                ws.cell(summary_row, 2).value = round(sum(vals) / len(vals), 2)
            else:
                total_dur = sum(r["duration_s"] for r in rows) / 60
                ws.cell(summary_row, 2).value = round(total_dur, 1)

    # Auto-filter on header row
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{1 + n}"

    # ─── Per-speaker sheet ──────────────────────────────────────
    if per_speaker_rows:
        ws2 = wb.create_sheet("Per speaker")
        spk_columns = [
            ("file", 50),
            ("folder", 30),
            ("speaker_id", 14),
            ("role", 12),
            ("talk_time_s", 12),
            ("n_turns", 9),
            ("avg_turn_s", 11),
            ("speaker_snr_db", 12),
            ("speaker_rms_dbfs", 14),
            ("speaker_clipping_pct", 18),
            ("quality", 13),
            ("score", 8),
            ("comment", 60),
        ]
        spk_headers = [c[0] for c in spk_columns]
        ws2.append(spk_headers)
        for i, h in enumerate(spk_headers, 1):
            ws2.cell(1, i).font = Font(bold=True)
            ws2.cell(1, i).alignment = Alignment(horizontal="center")
            ws2.column_dimensions[get_column_letter(i)].width = spk_columns[i-1][1]
        ws2.freeze_panes = "A2"

        role_colors = {
            "ENUMERATOR": "BDD7EE",  # light blue
            "RESPONDENT": "F8CBAD",  # light orange
            "OTHER":      "D9D9D9",  # light gray
            "UNKNOWN":    "BFBFBF",  # gray
        }

        for r in per_speaker_rows:
            ws2.append([
                r["file"], r["folder"],
                r["speaker_id"], r["role"],
                r["talk_time_s"], r["n_turns"], r["avg_turn_s"],
                round(r["speaker_snr_db"], 1),
                round(r["speaker_rms_dbfs"], 1),
                round(r["speaker_clipping_pct"], 3),
                r["quality"], r["score"], r["comment"],
            ])
            # Color the role cell
            role_idx = spk_headers.index("role") + 1
            cell = ws2.cell(row=ws2.max_row, column=role_idx)
            color = role_colors.get(r["role"])
            if color:
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
            # Color quality cell
            q_idx = spk_headers.index("quality") + 1
            qcell = ws2.cell(row=ws2.max_row, column=q_idx)
            color = quality_colors.get(r["quality"])
            if color:
                qcell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                qcell.font = Font(bold=True)
                qcell.alignment = Alignment(horizontal="center")

        ws2.auto_filter.ref = f"A1:{get_column_letter(len(spk_headers))}{1 + len(per_speaker_rows)}"

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


# ─── Main ────────────────────────────────────────────────────────────────

def main():
    p = argparse.ArgumentParser(
        description="Simple audio quality assessment to Excel.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    p.add_argument("folder", type=Path,
                   help="Folder containing audio files (searched recursively)")
    p.add_argument("-o", "--output", type=Path, default=None,
                   help="Output Excel path (default: <folder>/quality_report.xlsx)")
    p.add_argument("--extensions", nargs="+",
                   default=[".m4a", ".wav", ".mp3", ".flac", ".ogg", ".opus", ".aac"],
                   help="Audio file extensions to scan")
    p.add_argument("--max-files", type=int, default=-1,
                   help="Cap number of files to process (-1 = all)")
    p.add_argument("--with-speakers", action="store_true",
                   help="Also run speaker diarization and add a 'Per speaker' "
                        "sheet with ENUMERATOR/RESPONDENT roles. "
                        "Slower (~10-60s/file) and needs HF token + pyannote.audio.")
    p.add_argument("--max-speaker-files", type=int, default=-1,
                   help="When --with-speakers, cap diarization to first N usable "
                        "files (-1 = all). Useful for sanity-checking on large sets.")
    p.add_argument("--diarize-all", action="store_true",
                   help="Diarize every file, even those flagged unusable "
                        "(BAD/POOR/CLIPPED/MUTED/BROKEN). Default skips them.")
    args = p.parse_args()

    if not args.folder.exists():
        print(f"ERROR: folder not found: {args.folder}")
        sys.exit(1)

    if args.output is None:
        args.output = args.folder / "quality_report.xlsx"

    # Find audio files
    exts = {e.lower() if e.startswith(".") else f".{e.lower()}"
            for e in args.extensions}
    audio_files = []
    for f in sorted(args.folder.rglob("*")):
        if f.is_file() and f.suffix.lower() in exts:
            audio_files.append(f)
            if args.max_files > 0 and len(audio_files) >= args.max_files:
                break

    if not audio_files:
        print(f"No audio files found in {args.folder} "
              f"(searched for: {sorted(exts)})")
        sys.exit(0)

    print(f"Found {len(audio_files)} audio files in {args.folder}")
    print(f"Output: {args.output}")

    # Optionally load diarizer
    diarizer = None
    if args.with_speakers:
        print(f"\nLoading speaker diarizer (pyannote 3.1)...")
        diarizer = load_diarizer()
        if diarizer is None:
            print("⚠ Diarizer failed to load — proceeding without per-speaker analysis")
        else:
            print("  [OK] Diarizer ready")

    # Process each file
    import time
    rows = []
    per_speaker_rows = []
    n_errors = 0
    n_diarized = 0
    n_skipped_diar = 0
    SKIP_FLAGS = {"BAD", "POOR", "CLIPPED", "MUTED", "BROKEN", "EMPTY"}
    t0 = time.time()
    print()
    for i, path in enumerate(audio_files, 1):
        try:
            wav, sr = load_audio_mono16k(path)
            metrics = compute_metrics(wav, sr)
            flag, comment, score = classify_quality(
                metrics["snr_db"], metrics["speech_pct"],
                metrics["clipping_pct"], metrics["rms_dbfs"],
                metrics["duration_s"],
            )
            rel_path = path.relative_to(args.folder)
            row = {
                "file": str(rel_path.name),
                "folder": str(rel_path.parent) if rel_path.parent != Path(".") else "",
                "quality": flag,
                "score": score,
                "comment": comment,
                **metrics,
            }
            rows.append(row)

            # Per-speaker analysis (optional)
            speakers_info = ""
            if diarizer is not None:
                # Skip unusable audio unless user asked otherwise
                if not args.diarize_all and flag in SKIP_FLAGS:
                    n_skipped_diar += 1
                    speakers_info = f"  diar skipped ({flag})"
                elif (args.max_speaker_files > 0
                      and n_diarized >= args.max_speaker_files):
                    speakers_info = f"  diar skipped (cap {args.max_speaker_files})"
                else:
                    t_diar = time.time()
                    spk_rows = per_speaker_analysis(wav, diarizer, sr)
                    diar_elapsed = time.time() - t_diar
                    n_diarized += 1
                    for sr_row in spk_rows:
                        sr_row["file"] = row["file"]
                        sr_row["folder"] = row["folder"]
                        per_speaker_rows.append(sr_row)
                    spk_summary = ", ".join(
                        f"{r['role']}({r['quality']})" for r in spk_rows
                    ) if spk_rows else "no speakers"
                    speakers_info = f"  speakers ({diar_elapsed:.1f}s): {spk_summary}"

            # Print short progress (every file in diarization mode since each is slow)
            verbose = diarizer is not None
            if verbose or i <= 5 or i % 10 == 0 or i == len(audio_files):
                elapsed = time.time() - t0
                rate = i / max(elapsed, 1e-6)
                eta = (len(audio_files) - i) / max(rate, 1e-6)
                eta_str = f"({rate:.1f}/s, ETA {eta:.0f}s)" if i < len(audio_files) else ""
                print(f"  [{i:>3}/{len(audio_files)}] {flag:<10} "
                      f"SNR={metrics['snr_db']:>5.1f}dB  "
                      f"speech={metrics['speech_pct']:>4.0f}%  "
                      f"{rel_path.name[:50]} {eta_str}")
                if speakers_info:
                    print(f"      {speakers_info}")
        except Exception as e:
            n_errors += 1
            print(f"  [{i:>3}/{len(audio_files)}] ERROR ({type(e).__name__}): {path.name}")
            continue

    if diarizer is not None:
        print(f"\n  Diarized: {n_diarized} files, skipped: {n_skipped_diar} (unusable)")

    # Write Excel
    print(f"\nWriting Excel report -> {args.output}")
    write_excel(rows, args.output,
                per_speaker_rows=per_speaker_rows if per_speaker_rows else None)

    # Print summary to stdout
    from collections import Counter
    print(f"\n{'=' * 60}")
    print(f"SUMMARY: {len(rows)} files processed ({n_errors} errors)")
    print(f"{'=' * 60}")
    quality_counts = Counter(r["quality"] for r in rows)
    for flag in ["EXCELLENT", "GOOD", "FAIR", "LOW SPEECH", "EMPTY",
                 "POOR", "BAD", "CLIPPED", "MUTED", "BROKEN"]:
        count = quality_counts.get(flag, 0)
        if count > 0:
            pct = 100 * count / len(rows)
            print(f"  {flag:<12} {count:>4}  ({pct:.0f}%)")

    if rows:
        avg_snr = sum(r["snr_db"] for r in rows) / len(rows)
        avg_speech = sum(r["speech_pct"] for r in rows) / len(rows)
        total_dur = sum(r["duration_s"] for r in rows) / 60
        print(f"\n  Mean SNR:       {avg_snr:.1f} dB")
        print(f"  Mean speech %:  {avg_speech:.1f} %")
        print(f"  Total duration: {total_dur:.1f} min")

    print(f"\n  [OK] Open: {args.output}")


if __name__ == "__main__":
    main()
