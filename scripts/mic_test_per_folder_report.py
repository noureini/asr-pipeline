"""
Per-folder audio quality report.

Reads a `mic-test-report.json` produced by `asr-pipeline test-mics`
and aggregates the per-file acoustic metrics into a per-folder
(interview-key) summary. Outputs:

    <out_dir>/per-folder-report.csv   - sortable spreadsheet
    <out_dir>/per-folder-report.png   - rendered table image

Usage:
    uv run python scripts/mic_test_per_folder_report.py \
        outputs/bangladesh_phone_mic_test/mic-test-report.json
"""
from __future__ import annotations

import argparse
import json
import statistics as stats
from datetime import datetime
from pathlib import Path

import matplotlib.pyplot as plt


def classify_quality(snr: float, speech_pct: float) -> tuple[str, str]:
    """Return a (flag, comment) pair summarising recording quality.

    Thresholds chosen for phone-mic field interview audio.
    """
    if snr < 8:
        return ("BAD", "near-silent / no usable speech")
    if snr < 12:
        return ("POOR", "very noisy, expect high WER")
    if snr >= 20 and speech_pct >= 40:
        return ("GOOD", "clean, speech-rich")
    if snr >= 15 and speech_pct < 25:
        return ("LOW SPEECH", "high SNR but mostly silent — mic likely far / muted")
    if snr >= 12 and speech_pct >= 30:
        return ("OK", "usable, moderate noise")
    return ("OK", "borderline")


def aggregate_by_folder(report: dict) -> list[dict]:
    """Collapse all per-file metrics into one row per folder_key."""
    rows: dict[str, list[dict]] = {}
    for summary in report.get("mic_summaries", []):
        for f in summary.get("files", []):
            rows.setdefault(f["folder_key"], []).append(f)

    out = []
    for folder_key in sorted(rows):
        files = rows[folder_key]

        def _mean(field: str) -> float:
            return stats.fmean(float(f[field]) for f in files)

        def _sum(field: str) -> int:
            return sum(int(f[field]) for f in files)

        snr = _mean("snr_db")
        speech_pct = _mean("speech_ratio") * 100.0
        flag, comment = classify_quality(snr, speech_pct)
        out.append({
            "folder": folder_key,
            "mic": files[0]["mic_name"],
            "n_files": len(files),
            "duration_s": sum(float(f["duration_s"]) for f in files),
            "snr_db": snr,
            "clip_pct": _mean("clipping_ratio") * 100.0,
            "plosives": _sum("plosive_spike_count"),
            "rolloff_hz": _mean("spectral_rolloff_hz"),
            "bandwidth_hz": _mean("effective_bandwidth_hz"),
            "crosstalk": _mean("crosstalk_ratio"),
            "rms_dbfs": _mean("rms_dbfs"),
            "speech_pct": speech_pct,
            "quality_flag": flag,
            "comment": comment,
        })
    return out


def extract_per_file(report: dict) -> list[dict]:
    """Flatten the report into one row per audio file, with the same
    quality classification used for the per-folder summary."""
    out: list[dict] = []
    for summary in report.get("mic_summaries", []):
        for f in summary.get("files", []):
            snr = float(f["snr_db"])
            speech_pct = float(f["speech_ratio"]) * 100.0
            flag, comment = classify_quality(snr, speech_pct)
            out.append({
                "folder": f["folder_key"],
                "filename": Path(f["file_path"]).name,
                "mic": f["mic_name"],
                "duration_s": float(f["duration_s"]),
                "snr_db": snr,
                "clip_pct": float(f["clipping_ratio"]) * 100.0,
                "plosives": int(f["plosive_spike_count"]),
                "rolloff_hz": float(f["spectral_rolloff_hz"]),
                "bandwidth_hz": float(f["effective_bandwidth_hz"]),
                "crosstalk": float(f["crosstalk_ratio"]),
                "rms_dbfs": float(f["rms_dbfs"]),
                "speech_pct": speech_pct,
                "quality_flag": flag,
                "comment": comment,
            })
    out.sort(key=lambda r: (r["folder"], r["filename"]))
    return out


def write_csv(rows: list[dict], path: Path) -> None:
    import csv
    with path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


FLAG_FILLS_HEX = {
    "GOOD": "C8F0C8",
    "OK": "FDF3C2",
    "LOW SPEECH": "FDE2B3",
    "POOR": "F7C4C4",
    "BAD": "E89696",
}


def _populate_sheet(
    ws,
    rows: list[dict],
    headers_display: list[str],
    keys: list[str],
    quality_col_idx: int,
    title: str,
) -> None:
    """Fill an openpyxl worksheet with title, header, data, and footer."""
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    ncols = len(headers_display)

    # Title row
    ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=13)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)

    # Header row
    header_fill = PatternFill("solid", fgColor="1F3B6E")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, name in enumerate(headers_display, start=1):
        c = ws.cell(row=2, column=col_idx, value=name)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    # Data rows
    flag_fills = {k: PatternFill("solid", fgColor=v) for k, v in FLAG_FILLS_HEX.items()}
    for row_idx, r in enumerate(rows, start=3):
        for col_idx, key in enumerate(keys, start=1):
            val = r[key]
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if isinstance(val, float):
                if key in ("clip_pct", "crosstalk"):
                    cell.number_format = "0.0000"
                else:
                    cell.number_format = "0.0"
        flag = r["quality_flag"]
        if flag in flag_fills:
            qcell = ws.cell(row=row_idx, column=quality_col_idx)
            qcell.fill = flag_fills[flag]
            qcell.font = Font(bold=True)

    # Footer with means
    last_data_row = 2 + len(rows)
    footer_row = last_data_row + 1
    ws.cell(row=footer_row, column=1, value="MEAN").font = Font(bold=True)
    skip_keys = {"folder", "filename", "mic", "speaker_id", "role",
                 "quality_flag", "comment"}
    for col_idx, key in enumerate(keys, start=1):
        if key in skip_keys:
            continue
        col_letter = get_column_letter(col_idx)
        formula = f"=AVERAGE({col_letter}3:{col_letter}{last_data_row})"
        cell = ws.cell(row=footer_row, column=col_idx, value=formula)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E8EEF9")
        if key in ("clip_pct", "crosstalk"):
            cell.number_format = "0.0000"
        else:
            cell.number_format = "0.0"

    # Freeze + filter + autosize
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(ncols)}{last_data_row}"
    for col_idx in range(1, ncols + 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            (len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(2, footer_row + 1)),
            default=10,
        )
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 50)


def write_xlsx(
    folder_rows: list[dict],
    file_rows: list[dict],
    path: Path,
    title: str,
    speaker_rows: list[dict] | None = None,
) -> None:
    """Write a formatted .xlsx with up to three sheets:
    'Per folder', 'Per file' and (optionally) 'Per speaker'.
    """
    from openpyxl import Workbook

    folder_headers = [
        "Folder ID", "Mic", "Files", "Duration (s)",
        "SNR (dB)", "Clip %", "Plosives",
        "Rolloff (Hz)", "Bandwidth (Hz)", "Crosstalk",
        "RMS (dBFS)", "Speech %", "Quality", "Comment",
    ]
    folder_keys = [
        "folder", "mic", "n_files", "duration_s",
        "snr_db", "clip_pct", "plosives",
        "rolloff_hz", "bandwidth_hz", "crosstalk",
        "rms_dbfs", "speech_pct", "quality_flag", "comment",
    ]
    file_headers = [
        "Folder ID", "Filename", "Mic", "Duration (s)",
        "SNR (dB)", "Clip %", "Plosives",
        "Rolloff (Hz)", "Bandwidth (Hz)", "Crosstalk",
        "RMS (dBFS)", "Speech %", "Quality", "Comment",
    ]
    file_keys = [
        "folder", "filename", "mic", "duration_s",
        "snr_db", "clip_pct", "plosives",
        "rolloff_hz", "bandwidth_hz", "crosstalk",
        "rms_dbfs", "speech_pct", "quality_flag", "comment",
    ]
    speaker_headers = [
        "Folder ID", "Filename", "Mic", "Speaker", "Role",
        "Talk time (s)", "Turns", "Avg turn (s)",
        "SNR (dB)", "Clip %", "Plosives",
        "Rolloff (Hz)", "Bandwidth (Hz)", "Crosstalk",
        "RMS (dBFS)", "Quality", "Comment",
    ]
    speaker_data_keys = [
        "folder", "filename", "mic", "speaker_id", "role",
        "talk_time_s", "n_turns", "avg_turn_s",
        "snr_db", "clip_pct", "plosives",
        "rolloff_hz", "bandwidth_hz", "crosstalk",
        "rms_dbfs", "quality_flag", "comment",
    ]

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Per folder"
    _populate_sheet(ws1, folder_rows, folder_headers, folder_keys,
                    quality_col_idx=13, title=f"{title} — per folder")

    ws2 = wb.create_sheet("Per file")
    file_title = title.replace("by Folder", "by File")
    _populate_sheet(ws2, file_rows, file_headers, file_keys,
                    quality_col_idx=13, title=f"{file_title} — per file")

    if speaker_rows:
        ws3 = wb.create_sheet("Per speaker")
        spk_title = title.replace("by Folder", "by Speaker")
        _populate_sheet(
            ws3, speaker_rows, speaker_headers, speaker_data_keys,
            quality_col_idx=16, title=f"{spk_title} — per speaker",
        )

    wb.save(path)


def render_png(rows: list[dict], path: Path, title: str) -> None:
    headers = [
        "Folder ID", "Mic", "Files", "Dur (s)",
        "SNR (dB)", "Clip %", "Plosives",
        "Rolloff (Hz)", "BW (Hz)", "Crosstalk",
        "RMS (dBFS)", "Speech %", "Quality", "Comment",
    ]

    def _fmt(r: dict) -> list[str]:
        return [
            r["folder"], r["mic"], str(r["n_files"]), f"{r['duration_s']:.1f}",
            f"{r['snr_db']:.1f}", f"{r['clip_pct']:.4f}", str(r["plosives"]),
            f"{r['rolloff_hz']:.0f}", f"{r['bandwidth_hz']:.0f}",
            f"{r['crosstalk']:.4f}", f"{r['rms_dbfs']:.1f}",
            f"{r['speech_pct']:.1f}", r["quality_flag"], r["comment"],
        ]

    table_data = [_fmt(r) for r in rows]

    # Footer: overall mean / median
    def _col_mean(key: str) -> float:
        return stats.fmean(r[key] for r in rows)

    # Tally quality flags for the footer
    from collections import Counter
    flag_counts = Counter(r["quality_flag"] for r in rows)
    flag_summary = " | ".join(
        f"{k}:{v}" for k, v in sorted(flag_counts.items())
    )

    footer_mean = [
        "MEAN", "—", f"{stats.fmean(r['n_files'] for r in rows):.1f}",
        f"{stats.fmean(r['duration_s'] for r in rows):.1f}",
        f"{_col_mean('snr_db'):.1f}", f"{_col_mean('clip_pct'):.4f}",
        f"{stats.fmean(r['plosives'] for r in rows):.1f}",
        f"{_col_mean('rolloff_hz'):.0f}", f"{_col_mean('bandwidth_hz'):.0f}",
        f"{_col_mean('crosstalk'):.4f}", f"{_col_mean('rms_dbfs'):.1f}",
        f"{_col_mean('speech_pct'):.1f}",
        "—", flag_summary,
    ]
    table_data.append(footer_mean)

    # Layout: scale figure with row count
    n_rows = len(table_data) + 1
    fig_height = max(3.0, 0.32 * n_rows + 1.2)
    fig, ax = plt.subplots(figsize=(15, fig_height))
    ax.axis("off")
    ax.set_title(title, fontsize=13, fontweight="bold", pad=12)

    table = ax.table(
        cellText=table_data,
        colLabels=headers,
        loc="center",
        cellLoc="center",
        colLoc="center",
    )
    table.auto_set_font_size(False)
    table.set_fontsize(8)
    table.scale(1.0, 1.25)

    # Header style
    for col_idx in range(len(headers)):
        cell = table[0, col_idx]
        cell.set_facecolor("#1f3b6e")
        cell.set_text_props(color="white", weight="bold")

    # Footer (mean) row style
    last_row = len(table_data)  # because header is row 0
    for col_idx in range(len(headers)):
        cell = table[last_row, col_idx]
        cell.set_facecolor("#e8eef9")
        cell.set_text_props(weight="bold")

    # Highlight best/worst SNR among data rows (col index 4)
    snr_vals = [r["snr_db"] for r in rows]
    best_snr, worst_snr = max(snr_vals), min(snr_vals)
    for i, r in enumerate(rows, start=1):
        if r["snr_db"] == best_snr:
            table[i, 4].set_facecolor("#d6f5d6")
        elif r["snr_db"] == worst_snr:
            table[i, 4].set_facecolor("#f9d6d6")

    # Color-code Quality column (index 12) per flag
    flag_colors = {
        "GOOD": "#c8f0c8",
        "OK": "#fdf3c2",
        "LOW SPEECH": "#fde2b3",
        "POOR": "#f7c4c4",
        "BAD": "#e89696",
    }
    for i, r in enumerate(rows, start=1):
        color = flag_colors.get(r["quality_flag"], "white")
        cell = table[i, 12]
        cell.set_facecolor(color)
        cell.set_text_props(weight="bold")

    fig.tight_layout()
    fig.savefig(path, dpi=180, bbox_inches="tight")
    plt.close(fig)


def main() -> None:
    p = argparse.ArgumentParser()
    p.add_argument("report_json", type=Path,
                   help="Path to mic-test-report.json")
    p.add_argument("-o", "--output-dir", type=Path, default=None,
                   help="Output directory (default: alongside the JSON)")
    args = p.parse_args()

    report = json.loads(args.report_json.read_text(encoding="utf-8"))
    rows = aggregate_by_folder(report)
    if not rows:
        raise SystemExit("No per-file metrics found in report.")

    out_dir = args.output_dir or args.report_json.parent
    out_dir.mkdir(parents=True, exist_ok=True)

    file_rows = extract_per_file(report)

    # Optional per-speaker JSON (produced by mic_test_per_speaker.py)
    speaker_rows: list[dict] = []
    speaker_json = out_dir / "per-speaker-report.json"
    if speaker_json.exists():
        try:
            speaker_rows = json.loads(speaker_json.read_text(encoding="utf-8")).get(
                "rows", []
            )
        except Exception as e:
            print(f"Warning: could not read {speaker_json}: {e}")

    folder_csv = out_dir / "per-folder-report.csv"
    file_csv = out_dir / "per-file-report.csv"
    png_path = out_dir / "per-folder-report.png"
    xlsx_path = out_dir / "per-folder-report.xlsx"

    # Excel may be open in another app, locking the file. Fall back to a
    # timestamped filename if the primary one isn't writable.
    def _writable(p: Path) -> bool:
        try:
            with p.open("a"):
                return True
        except Exception:
            return False
    if xlsx_path.exists() and not _writable(xlsx_path):
        ts = datetime.now().strftime("%H%M%S")
        xlsx_path = out_dir / f"per-folder-report-{ts}.xlsx"
        print(f"Note: original xlsx is locked; writing to {xlsx_path.name}")

    title = (
        f"Audio Quality by Folder — {report.get('language', '?')} — "
        f"{report.get('test_date', '?')[:10]} ({len(rows)} folders)"
    )
    write_csv(rows, folder_csv)
    write_csv(file_rows, file_csv)
    render_png(rows, png_path, title)
    write_xlsx(rows, file_rows, xlsx_path, title, speaker_rows=speaker_rows)

    sheets = "'Per folder', 'Per file'"
    if speaker_rows:
        sheets += ", 'Per speaker'"
    print(f"Wrote {folder_csv}")
    print(f"Wrote {file_csv}")
    print(f"Wrote {png_path}")
    print(f"Wrote {xlsx_path}  (sheets: {sheets})")
    if speaker_rows:
        print(f"  Per-speaker rows picked up from {speaker_json.name}: "
              f"{len(speaker_rows)}")


if __name__ == "__main__":
    main()
